# Import all relevant modules
import gurobipy as gp
import pandas as pd
from gurobipy import GRB
import sys
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time



# Load the workbook

wb = load_workbook('ModellingData.xlsx')
sheet1 = wb['Constants']

selector = int(input(" 1. UK Summer \n 2. UK Winter \n 3. GE Summer \n 4. GE Winter \n 5. SP Summer \n 6. SP Winter \n Enter number corresponding to selection: "))

if selector == 1:
    interval = float(sheet1['C21'].value)
    nTrades = int(sheet1['C24'].value)
    sheet2 = wb['UK_Summer']
    print("\nUK_Summer data loaded")
elif selector == 2: 
    interval = float(sheet1['C21'].value)
    nTrades = int(sheet1['C24'].value)
    sheet2 = wb['UK_Winter']
    print("\nUK_Winter data loaded")

elif selector == 3:
    interval = float(sheet1['C22'].value)
    nTrades = int(sheet1['C25'].value)
    sheet2 = wb['GE_Summer']
    print("\nGE_Summer data loaded")
elif selector == 4: 
    interval = float(sheet1['C22'].value)
    nTrades = int(sheet1['C25'].value)
    sheet2 = wb['GE_Winter']
    print("\nGE_Winter data loaded")

elif selector == 5: 
    interval = float(sheet1['C23'].value)
    nTrades = int(sheet1['C26'].value)
    sheet2 = wb['SP_Summer']
    print("\nSP_Summer data loaded")
elif selector == 6: 
    interval = float(sheet1['C23'].value)
    nTrades = int(sheet1['C26'].value)
    sheet2 = wb['SP_Winter']
    print("\nSP_Winter data loaded")
else:
    print("Invalid input")

nDays = int(input("Enter number of days to optimize: "))
#Initialise model constants

battCapacity = float(sheet1['C2'].value)
cycleLimit = float(sheet1['C3'].value)
pPVMax = float(sheet1['C4'].value)
pLoadMax = float(sheet1['C5'].value)
pImpMax = float(sheet1['C6'].value)
pExpMax = float(sheet1['C7'].value)
effPV = float(sheet1['C8'].value)
effCharg = float(sheet1['C9'].value)
effDisch = float(sheet1['C10'].value)
copHeat = float(sheet1['C11'].value)
copCool = float(sheet1['C12'].value)
pHeatMax = float(sheet1['C13'].value)
pCoolMax = float(sheet1['C14'].value)
thermalMass = float(sheet1['C15'].value)
thermalRes = float(sheet1['C16'].value)
tIntMin = float(sheet1['C17'].value)
tIntMax = float(sheet1['C18'].value)
pChargMax = float(sheet1['C19'].value)
pDischMax = float(sheet1['C20'].value)


# Create a new model
m = gp.Model("EnergyTrading")

# Create continuous model variables
pCool = m.addVar(lb=0, name="pCool", vtype=GRB.CONTINUOUS)
pHeat = m.addVar(lb=0, name="pHeat", vtype=GRB.CONTINUOUS)
pCharg = m.addVar(lb=0, name="pCharg", vtype=GRB.CONTINUOUS)
pDisch = m.addVar(lb=0, name="pDisch", vtype=GRB.CONTINUOUS)
pImp = m.addVar( lb=0, name="pImp", vtype=GRB.CONTINUOUS)
pExp = m.addVar(lb=0, name="pExp", vtype=GRB.CONTINUOUS)
pPV = m.addVar(lb=0, name="pPV", vtype=GRB.CONTINUOUS)
tInt = m.addVar(ub=tIntMax, lb=tIntMin, name="tInt", vtype=GRB.CONTINUOUS)
battLevel = m.addVar(ub=battCapacity, lb=0, name="battLevel", vtype=GRB.CONTINUOUS)

# Create binary decision variables
is_idm_import_mode = m.addVar(vtype=GRB.BINARY, name="is_idm_import_mode")
is_battery_charge_mode = m.addVar(vtype=GRB.BINARY, name="is_battery_charge_mode")
is_ehp_heat_mode = m.addVar(vtype=GRB.BINARY, name="is_ehp_heat_mode")


for i in range (nDays):
    for j in range(nTrades):
        
        curRow = (i*nTrades) + (j+3)
        
        # Update dynamic variables in the model
        price = float(sheet2.cell(row=curRow,column=3).value)
        priceOld = float(sheet2.cell(row=curRow-1,column=3).value)
        pPVA = float(sheet2.cell(row=curRow,column=4).value)
        tAmb = float(sheet2.cell(row=curRow,column=5).value)
        pLoad = float(sheet2.cell(row=curRow,column=6).value)
        tAmbOld = float(sheet2.cell(row=curRow-1,column=5).value)
        pHeatOld = float(sheet2.cell(row=curRow-1,column=8).value)
        pCoolOld = float(sheet2.cell(row=curRow-1,column=7).value)
        pChargOld = float(sheet2.cell(row=curRow-1,column=9).value)
        pDischOld = float(sheet2.cell(row=curRow-1,column=10).value)
        tIntOld = float(sheet2.cell(row=curRow-1,column=14).value)
        battLevelOld = float(sheet2.cell(row=curRow-1,column=15).value)

        # Constraint on operation of EHP
        m.addConstr(tInt - interval*((tAmbOld-tIntOld)/(thermalMass*thermalRes)+((copHeat*pHeatOld)-(copCool*pCoolOld))/thermalMass), GRB.EQUAL, tIntOld,"c0")

        # Constraints on energy storage 
        m.addConstr(battLevel - interval*((effCharg*pChargOld)-(pDischOld/effDisch)), GRB.EQUAL, battLevelOld,"c1")

        # Power flow constraint
        m.addConstr((pPV + pDisch + pImp - pHeat - pCool - pCharg - pExp), GRB.EQUAL, pLoad, "c2")
        m.addConstr(pPV <= pPVA*effPV, "c3")

       # Binary decision constraints
        m.addConstr(pExp <= pExpMax*(1-is_idm_import_mode),"c4")
        m.addConstr(pImp <= pImpMax*is_idm_import_mode, "c5")
        m.addConstr(pHeat <= pHeatMax*is_ehp_heat_mode, "c6")
        m.addConstr(pCool <= pCoolMax*(1-is_ehp_heat_mode), "c7")
        m.addConstr(pCharg <= pChargMax*is_battery_charge_mode, "c8")
        m.addConstr(pDisch <= pDischMax*(1-is_battery_charge_mode), "c9")     
        # Include simple heuristics
        
        if(pPVA>0.5*pPVMax or price<priceOld):
            m.addConstr(pCharg >= pChargMax*(1-battLevelOld/battCapacity), "h1")
        
        # Required heating/cooling power

        if(tAmb<tIntOld):
            m.addConstr(pHeat >= pHeatMax*(tIntMax-tIntOld)/tIntMax, "h4")
        else:
            m.addConstr(pCool >= pCoolMax*(tIntOld-tIntMin)/tIntMin, "h5")
        
        m.update() 

        # Set objective function
        obj = (price*pExp) - (price*pImp)
        m.setObjective(obj, GRB.MAXIMIZE)
        m.update()   
        # Run optimization
        
        m.optimize()
        # m.update()   
    
        # Update Spreadsheet with optimal solutions for each trading period
           
        sheet2.cell(row=curRow,column=14).value = m.getVarByName("tInt").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=15).value = m.getVarByName("battLevel").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=16).value = m.getObjective().getValue() 
        sheet2.cell(row=curRow,column=11).value = m.getVarByName("pImp").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=12).value = m.getVarByName("pExp").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=8).value = m.getVarByName("pHeat").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=7).value = m.getVarByName("pCool").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=13).value = m.getVarByName("pPV").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=9).value = m.getVarByName("pCharg").getAttr(GRB.Attr.X)
        sheet2.cell(row=curRow,column=10).value = m.getVarByName("pDisch").getAttr(GRB.Attr.X)

        wb.save("ModellingData.xlsx")
        print("\nTrading interval no. "+str(curRow-2)+" successfully optimized.\n")
        m.remove(m.getConstrs())
        # m.reset(0)
        # time.sleep(1)

print("\nFull Optimization completed.\nResults saved to worksheet.\n")
    
